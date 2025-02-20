import configparser
from cryptography.fernet import Fernet
import sys
from typing import List

import pyodbc # ODBC
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill



from PySide6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QMessageBox, QTableWidget, \
    QTableWidgetItem, QPlainTextEdit, QPushButton, QApplication
from PySide6.QtCore import QTimer, QRect
from Qt.main_window import Ui_MainWindow
from Qt.history_window import Ui_Form  # Importiere das generierte UI-Design
from decimal import Decimal

import json
import subprocess
import os
from datetime import datetime
from PyQt5.QtCore import Qt


#######################################################################################################################
# Releasenotes
#
# 30.06.24  0.1 Erste Version
# 15.07.24  0.2 Aufruf von mehreren Instanzen. Mit diesem Stand kann nun aufgebaut werden
# 17.11.24  0.3 Checkboxen welche HV-Version und Test und Prod in der Maske eingebaut
#               Initialwerte für die Checkboxen in der ini-Datei ablegen
# 20.11.24  0.4 Die Instanzliste mit der HV-Nummer und TST oder PROD erweitert
#               Sortfunktion für die Instanzliste eingebaut.
#               Berücksichtigung der CheckBoxen in bei der Abfrage
# 27.12.24  0.5 Das Festlegen des Table-Views erfolgt nun immer und nicht nur bei der ersten Instanz
#               Wenn die erste Instanz keine Daten hatte sind keine Werte ausgegeben worden.
#           0.6 History Funktion für die SQL-Statements
# 24.01.25  0.7 Instanzen, User und Passwort aus Dateien lesen
# 25.01.25  1.0 Alle Instanzen kommen nun aus der INI-Datei und auch die User und Passwörter
# 29.01.25  1.1 Ausgabe der Ergebnisse zusätzlich in eine Excel-Datei
# 15.02.25  1.2 Neue Class mit HistoryWindow qForms mit QT-Designer erstellt und eingebaut
#               in diesem Fenster die History-Einträge ausgeben und auswähle
#               NewWindow in SQLWindow geändert
#               Default Instanz eingebaut
#               Mit der Open-Funktion aus dem Menü mit nodepad++ die Json datei öffnen
# 17.02.25      Sql-Aufruf mit try und except abgesichert
# 18.02.25      Hinzufügen des richtigen Pfads zur Tabelle optimiert
#               Tabelle kann nun mit form und join vorkommen. Bei Tabellen mit 2 oder 3
#               wird die COM-Bibliothek genommen
#               Menüpunkt zum Öffnen der Excel-Datei eingefügt
#
#
# --------------------Version
version = "1.2Branche-Statement-Histoy-mit-Anzeige"
# ---------------------------------

# Defaultwerte aus der Ini-Datei lesen
config = configparser.ConfigParser()
config.read('SQL-master.ini', encoding='utf-8')
HV9 = config.get('Einstellungen', 'HV9')
HV8 = config.get('Einstellungen', 'HV8')
HV7 = config.get('Einstellungen', 'HV7')
Nonhv = config.get('Einstellungen', 'NONHV')
Test = config.get('Einstellungen', 'TEST')
Prod = config.get('Einstellungen', 'PROD')
Default_instanz = config.get('Einstellungen', 'default_instanz')
Default_instanz_aktiv = config.get('Einstellungen', 'default_instanz_activ')


#Instanzen-Datei
instanzen_datei = "instanzen.ini"
password_ini_file = "sql-master-password.ini"



# Globale Variable
sql_input = "select count(*) from pfistam where fsfirm <> '' "
sql_commend =""
history_index = 0
header_zeilen = False
einstellung_uebernehmen = '0'
name_excel_datei = "sql_results.xlsx"

HISTORY_FILE_JSON = "sql_history.json"


# Neue Excel-Arbeitsmappe erstellen
workbook = Workbook()
sheet = workbook.active
sheet.title = "SQL Ergebnisse"


# Entschlüssele das Passwort
def decrypt_text(encrypted_text, key):
    """
    Entschlüsselt einen verschlüsselten Text mithilfe eines Verschlüsselungsschlüssels.

    Diese Funktion verwendet die Fernet-Verschlüsselung aus der `cryptography`-Bibliothek,
    um den verschlüsselten Text zu entschlüsseln. Der Schlüssel muss derselbe sein,
    der zur Verschlüsselung des Textes verwendet wurde.

    Args:
        encrypted_text (bytes): Der verschlüsselte Text als Byte-String.
        key (bytes): Der Schlüssel (Byte-String), der für die Fernet-Entschlüsselung verwendet wird.

    Returns:
        str: Der entschlüsselte Text in Klartext.

    """
    f = Fernet(key)
    decrypted_text = f.decrypt(encrypted_text).decode()
    return decrypted_text




def load_key():
    """
    Laden des Keyfiles
    Das Key-File liegt parallel zum Programm mit dem Namen secret.key

    Returns: Key für die Verschlüsselung / Entschlüsselung

    """
    return open("secret.key", "rb").read()


# Lese das verschlüsselte Passwort aus der INI-Datei
def read_decrypted_from_ini(section, key):
    """
    Lesen von verschlüsselten Texten aus einem INI-File

    Args:
        section (String): Section Name (hier der Maschinenname)
        key (String): User oder Password

    Returns (String): Text in lesbarer Form

    """
    config = configparser.ConfigParser()
    config.read(password_ini_file)

    # Überprüfen, ob der Abschnitt und der Schlüssel vorhanden sind
    if section in config and key in config[section]:
        encrypted_text = config[section][key]
        return encrypted_text
    else:
        print(f'Kein {key} in section {section} in der INI-Datei gefunden.')
        return ''

def is_sql_in_history(sql_statement):
    """Prüft ob ein sql-Statement schon in der history Datei enthalten ist"""
    try:
        with open(HISTORY_FILE_JSON, "r", encoding="utf-8") as file:
            history = json.load(file)

        # Überprüfe, ob das SQL-Statement bereits in der Historie ist
        for entry in history:
            if entry.get("sql") == sql_statement:
                return True  # SQL-Statement wurde gefunden

    except (FileNotFoundError, json.JSONDecodeError):
        return False  # Datei existiert nicht oder ist fehlerhaft

    return False  # SQL-Statement wurde nicht gefunden


def save_to_json_history(sql_statement, sql_commend):
    """Speichert ein SQL-Statement mit Metadaten in eine JSON-Datei."""
    history = []
    if os.path.exists(HISTORY_FILE_JSON):
        with open(HISTORY_FILE_JSON, "r") as file:
            history = json.load(file)
    entry = {"commend": sql_commend, "sql": sql_statement, "timestamp": datetime.now().isoformat()}
    history.append(entry)
    with open(HISTORY_FILE_JSON, "w") as file:
        json.dump(history, file, indent=4)

def read_json_history():
    """Liest die JSON-History und gibt die Einträge zurück."""
    if not os.path.exists(HISTORY_FILE_JSON):
        return []
    with open(HISTORY_FILE_JSON, "r") as file:
        return json.load(file)

def aufbereiten_instanz_liste():
    """
    Liest die Instanz-Datei, und die Datei mit dem User und das Passwort
    Entschlüsselt das Passwort und User und füllt die Instanz-Liste


    RETURN: gefüllte Instanzenliste
    """
    config = configparser.ConfigParser()
    config.optionxform = str  # Behalte die Groß-/Kleinschreibung der Schlüssel bei
    config.read(instanzen_datei)


    key = load_key()
    instanzen_liste = []
    # Alle Sektionen (Abschnitte mit den Maschinen) auslesen
    for maschine in config.sections():

        # Alle Key (Key mit HV-Typ) in dieser Sektion auslesen
        for hv_typ, value in config.items(maschine):
            values = value.split(",")  # Teilt den String in eine Liste

            for instanz in values:  # Schleife über die Werte in der Liste

                user = decrypt_text(read_decrypted_from_ini(maschine, "user").encode(), key)
                password = decrypt_text(read_decrypted_from_ini(maschine, "password").encode(), key)

                # dem HV-Typ aus dem Key noch in Werte teilen
                # HV9_TST   HV9   und   TST
                hv_typ_teil =hv_typ.split("_")
                instanzen_liste.append((maschine, user, password, instanz.strip(),
                                        hv_typ_teil[0], hv_typ_teil[1]))

    # Die resultierende Liste ausgeben (zu Testzwecken)
    #for entry in instanzen_liste:
    #    print(entry)

    return instanzen_liste




class MainWindow(QMainWindow):

    global sql_input
    def __init__(self):
        super(MainWindow, self).__init__()
        self.gui = Ui_MainWindow()
        self.gui.setupUi(self)

        # CheckBoxen vorbelegen
        if HV9 == '1':
            self.gui.checkBox_HV9.setChecked(True)
        if HV8 == '1':
            self.gui.checkBox_HV8.setChecked(True)
        if HV7 == '1':
            self.gui.checkBox_HV7.setChecked(True)
        if Nonhv == '1':
            self.gui.checkBox_NONHV.setChecked(True)
        if Test == '1':
            self.gui.checkBox_Test.setChecked(True)
        if Prod == '1':
            self.gui.checkBox_PROD.setChecked(True)
        if Default_instanz_aktiv == '1':
            self.gui.checkBox_default_instanz.setChecked(True)
        self.gui.lineEdit_default_instanz.setText(Default_instanz)


        # Verbinde Menüpunkte mit Methoden
        self.gui.actionOpen_History.triggered.connect(self.open_file)
        self.gui.actionOpen_Excel.triggered.connect(self.open_file_excel)
        self.gui.actionExit.triggered.connect(self.exit)
        self.gui.actionStart.triggered.connect(self.abfrage)
        self.gui.action_ber.triggered.connect(self.show_about)
        self.gui.actionSQL_Eingabe.triggered.connect(self.sql_eingabe)
        self.gui.actionSQL_History.triggered.connect(self.sql_history)

        # Verbinden der Check-Boxen mit einer Änderungs-Methode
        self.gui.checkBox_HV9.stateChanged.connect(self.on_checkbox_changed)
        self.gui.checkBox_HV8.stateChanged.connect(self.on_checkbox_changed)
        self.gui.checkBox_HV7.stateChanged.connect(self.on_checkbox_changed)
        self.gui.checkBox_NONHV.stateChanged.connect(self.on_checkbox_changed)
        self.gui.checkBox_Test.stateChanged.connect(self.on_checkbox_changed)
        self.gui.checkBox_PROD.stateChanged.connect(self.on_checkbox_changed)
        self.gui.checkBox_Einstellung.stateChanged.connect(self.on_checkbox_changed)
        self.gui.checkBox_default_instanz.stateChanged.connect(self.on_checkbox_changed)

        #Verbinden textEdit Änderung mit Änderungs-Methode
        self.gui.lineEdit_default_instanz.textChanged.connect(self.default_instance_changed)

    def default_instance_changed(self):
        global Default_instanz

        Default_instanz = self.gui.lineEdit_default_instanz.text()


    def on_checkbox_changed(self):
        global HV9, HV8, HV7, Nonhv, Test, Prod, einstellung_uebernehmen, Default_instanz_aktiv

        # zuerst alle globale Variable ausmachen
        HV9 = '0'
        HV8 = '0'
        HV7 = '0'
        Nonhv = '0'
        Test = '0'
        Prod = '0'
        einstellung_uebernehmen = '0'
        Default_instanz_aktiv = '0'

        # Die Werte der Checkboxen in die globalen Variablen wieder an machen die an sind
        if self.gui.checkBox_HV9.isChecked():
            HV9 = '1'
        if self.gui.checkBox_HV8.isChecked():
            HV8 = '1'
        if self.gui.checkBox_HV7.isChecked():
            HV7 = '1'
        if self.gui.checkBox_NONHV.isChecked():
            Nonhv = '1'
        if self.gui.checkBox_Test.isChecked():
            Test = '1'
        if self.gui.checkBox_PROD.isChecked():
            Prod = '1'
        if self.gui.checkBox_Einstellung.isChecked():
            einstellung_uebernehmen = '1'
        if self.gui.checkBox_default_instanz.isChecked():
            Default_instanz_aktiv = '1'






    def open_file(self):
        """Öffnet die Json-History Datei mit einem Editor"""
        print(read_json_history())
        file_path = HISTORY_FILE_JSON  # Pfad zur JSON-Datei

        if os.path.exists(file_path):  # Prüfen, ob die Datei existiert
            try:
                if os.name == "nt":  # Windows
                    os.startfile(file_path)
                elif os.name == "posix":  # Linux / macOS
                    subprocess.Popen(["xdg-open", file_path])
            except Exception as e:
                QMessageBox.warning(self, "Fehler", f"Konnte die Datei nicht öffnen:\n{str(e)}")
        else:
            QMessageBox.warning(self, "Fehler", "Die Datei sql_history.json wurde nicht gefunden!")

    def open_file_excel(self):
        """Öffnet die Excel Datei """

        file_path = HISTORY_FILE_JSON  # Pfad zur JSON-Datei

        if os.path.exists(file_path):  # Prüfen, ob die Datei existiert
            try:
                if os.name == "nt":  # Windows
                    os.startfile(name_excel_datei)
                elif os.name == "posix":  # Linux / macOS
                    subprocess.Popen(["xdg-open", name_excel_datei])
            except Exception as e:
                QMessageBox.warning(self, "Fehler", f"Konnte die Datei nicht öffnen:\n{str(e)}")
        else:
            QMessageBox.warning(self, "Fehler", "Die Datei "+name_excel_datei+" wurde nicht gefunden!")

    def show_about(self):
        QMessageBox.about(self, "About", "This is a PyQt5/PySide6 application")

    def show_help(self):
        self.gui.label.setText("Help clicked")



    def exit(self):
        exit()

    def sql_eingabe(self):
        try:
            self.table_widget.hide()
        except:
            print('hide ging schief')


        # self.gui.plainTextEdit.show()
        # text = self.gui.plainTextEdit.toPlainText()
        # print(text)

        # Erstelle und zeige ein neues Fenster mit QPlainTextEdit
        self.sql_window = SqlWindow(self)
        self.sql_window.show()

    def sql_history(self):
        print("SQL-History")

        try:
            self.table_widget.hide()
        except:
            print('hide ging schief')


        # self.gui.plainTextEdit.show()
        # text = self.gui.plainTextEdit.toPlainText()
        # print(text)

        # Erstelle und zeige ein neues Fenster mit QPlainTextEdit
        self.history_window = HistoryWindow(self)
        self.history_window.show()





    def abfrage(self):
        global sql_input, HV9, HV8, HV7, Nonhv, Test, Prod, header_zeilen, einstellung_uebernehmen
        print("Taste Start gedrückt")

        # Löscht alle Zeilen inkl. der Kopfzeile
        # Flag für header_zeilen auf False, damit der neue Header erstellt wird
        sheet.delete_rows(1, sheet.max_row)
        header_zeilen = False



        # Verhindern, dass beide Checkboxen für Test und Prod abgewählt sind
        if Test == '0'  and Prod == '0':
            # Wenn beide nicht ausgewählt sind, eine Test Checkbox wieder aktivieren
            sender = self.sender()
            self.gui.checkBox_Test.setChecked(True)
            Test = '1'




        # Speichern des Status der CheckBoxen wenn dies angewählt ist
        ## TO DO Einstellung muss noch mal geprüft werden
        if einstellung_uebernehmen == '1':
            config.set('Einstellungen', 'HV9', HV9)
            config.set('Einstellungen', 'HV8', HV8)
            config.set('Einstellungen', 'HV7', HV7)
            config.set('Einstellungen', 'Nonhv', Nonhv)
            config.set('Einstellungen', 'Test', Test)
            config.set('Einstellungen', 'Prod', Prod)
            config.set('Einstellungen', 'default_instanz_activ', Default_instanz_aktiv)
            config.set('Einstellungen', 'default_instanz', Default_instanz)

            # Schreiben der Änderungen zurück in die INI-Datei
            with open('SQL-master.ini', 'w') as configfile:
                config.write(configfile)


        # Haupt-Widget und Layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Erstelle ein QTableWidget
        self.table_widget = QTableWidget()



        instanz_list = aufbereiten_instanz_liste()



        self.zeilen_counter = 0

        print(f' sql_input aus der Eingabe {sql_input}')

        # Sortieren nach dem letzten und dann nach dem vorletzten Eintrag
        sorted_list = sorted(instanz_list, key=lambda x: (x[-1], x[-2]))




        zeile_in_for = 0
        for i, zeile in enumerate(sorted_list):


            # wenn der Tabellen name identifiziert wurde, die Instanz und die Bibl. noch erweitern

            # # Abfrage, ob die Instanz zu der gewünschten Auswahl gehört und ob Test oder Prod
            # # gewünscht ist. Wenn nicht wird in der for-Schleife der nächste Eintrag verwendet
            # if (zeile[4] == "HV9" and HV9 == '0'):
            #     continue
            # elif (zeile[4] == "HV8" and HV8 == '0'):
            #     continue
            # elif (zeile[4] == "HV7" and HV7 == '0'):
            #     continue
            # elif (zeile[5] == "TST" and Test == '0'):
            #     continue
            # elif (zeile[5] == "PROD" and Prod == '0'):
            #     continue


            # Dies ist die elegantere Variante um die Bedingung abzufragen um
            # die Schleife zu beenden
            # Bedingungen als Dictionary definieren
            conditions = {
                "HV9": HV9,
                "HV8": HV8,
                "HV7": HV7,
                "NONHV": Nonhv,
                "TST": Test,
                "PROD": Prod
            }
            # Überprüfen, ob die Zeile zu den Bedingungen passt
            if (zeile[4] in conditions and conditions[zeile[4]] == '0') or \
                    (zeile[5] in conditions and conditions[zeile[5]] == '0'):
                continue

            #neue Ermittlung des Pfades
            ##print(f'SQL mit dem Path {update_sql_with_paths(sql_input,zeile[3])}')
            sql_query =update_sql_with_paths(sql_input, zeile[3])

            ### Die alte Pfad-Ermittlung wird deaktiviert
            # if table_name:
            #     # Erstelle den neuen Tabellennamen mit der Instanz
            #     new_table_name = f"{zeile[3]}DATV7.{table_name}"
            #     # Ersetze den Tabellennamen im SQL-Statement
            #     sql_query = re.sub(rf'\b{table_name}\b', new_table_name, sql_input, flags=re.IGNORECASE)
            #     #print(f'SQL-statement nach der Aufbereitung: {sql_query}')


            self.sql_pro_instanz(sql_query,zeile[0],zeile[1],zeile[2], zeile[3], layout, zeile_in_for, zeile[4], zeile[5])

            #print(f'Ende der for schleife {i}')
            zeile_in_for += 1

        # Excel-Datei speichern
        try:
            workbook.save(name_excel_datei)
            print(f"Daten wurden in die Datei {name_excel_datei} geschrieben.")
        except:
            print(f"Datei {name_excel_datei} konnte nicht geschrieben werden.")






    # Methode um pro Instanz das SQL-Statement auszuführen am Bildschirm auszugeben
    def sql_pro_instanz(self, sql_statement, host, user, passwort, instanz, layout, nummer, version, art):

        global header_zeilen

        #wenn die default-Instanz aktiv ist, dann nur weiter machen, wenn die instanz auch der Default-Instanz
        #entspricht
        if Default_instanz_aktiv == '1' and Default_instanz != instanz:
            return

        print(f'Zeilen-Zähler  {instanz}: {self.zeilen_counter}')

        ## Wenn kein SQL auf der ISeries erfolgen soll die beiden Zeilen aktivieren
        ##print('********** zum Test keine SQL mit einer Verbindung ausführen')
        ##return

        # Erstelle die Verbindungszeichenfolge mit f-strings
        conn_str = (f'DRIVER={{IBM i Access ODBC Driver}};'
                    f'SYSTEM={host};'
                    f'UID={user};'
                    f'PWD={passwort}')



        # Stelle die Verbindung her
        try:
            conn = pyodbc.connect(conn_str)
        except pyodbc.Error as e:
            print(f'Fehler beim Herstellen der Verbindung: {e}')
            QMessageBox.critical(self, "Verbindungsfehler", f"Fehler beim Herstellen der Verbindung: {e}")

            return

        # Erstelle einen Cursor-Objekt
        cursor = conn.cursor()

        try:
            cursor.execute(sql_statement)
        except Exception as e:
            print(f"SQL-Statement fehlgeschlagen:\n{sql_statement}")
            print(f"Fehlermeldung: {e}")
            return


        # Hole alle Zeilen der Abfrage
        rows = cursor.fetchall()
        if len(rows) == 0:
            print("Keine Daten gefunden")
            return


        # Ab hier werden die Daten ausgegeben

        #***************Test für EXCEL-Ausgabe**********************'

        # Überschriften nur in der ersten Zeile
        if header_zeilen == False:
            # Spaltennamen aus der Abfrage holen
            column_names = ["Version"] + ["Art"] + ["Instanz"] +[description[0] for description in cursor.description]

            # Spaltenüberschriften in die erste Zeile schreiben
            sheet.append(column_names)
            header_zeilen = True
            #Autofilter in die Überschrift mit den Feldnamen setzen
            sheet.auto_filter.ref = sheet.dimensions

            # Erste Zeile Fixieren
            sheet.freeze_panes = "A2"

            # Formatierung der Kopfzeile Schrift in Bold und Hintergrund Grau
            header_font = Font(bold=True)  # Fett
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Grau
            for cell in sheet[1]:  # Erste Zeile (Kopfzeile)
                cell.font = header_font
                cell.fill = header_fill

        # Datenzeilen in das Excel-Blatt schreiben
        for row in rows:
            sheet.append([version]+[art]+[instanz]+list(row))  # Jede Zeile in eine Liste umwandeln

        # Spaltenbreite automatisch anpassen
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Spaltenbuchstabe (z. B. "A", "B", ...)
            for cell in column:
                if cell.value:  # Falls die Zelle nicht leer ist
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2  # Etwas zusätzlichen Platz hinzufügen
            sheet.column_dimensions[column_letter].width = adjusted_width


        #***********************************************************


        # Definieren der Breite und Länge der Tabelle für die Ausabe und Feldnamen setzen
        # Dies wird bei jeder Instanz gemacht, da nicht sicher ist dass bei der ersten Instanz
        # schon etwas gefunden wird.

        # Anzahl der Spalten festlegen
        #self.table_widget.setRowCount(len(rows))
        self.table_widget.setRowCount(999999)
        self.table_widget.setColumnCount(len(rows[0]) + 3)  # + 4 Felder für die Version, Art und Instanz

        layout.addWidget(self.table_widget)


        # Feldnamen aus den Metadaten der Abfrage abrufen
        column_names: list[str] = ["Version"] + ["Art"] + ["Instanz"] + [desc[0] for desc in cursor.description]  # Spalte 1,2 und ist für Version Art und Instanz
        # Feldnamen als Überschrift ausgeben
        self.table_widget.setHorizontalHeaderLabels(column_names)

        # Fülle die Tabelle mit Daten
        for row_idx, row in enumerate(rows):
            self.table_widget.setItem(row_idx+self.zeilen_counter, 0, QTableWidgetItem(version))  # Spalte 1 ist die Version)
            self.table_widget.setItem(row_idx + self.zeilen_counter, 1,QTableWidgetItem(art))  # Spalte 2 ist die Version)
            self.table_widget.setItem(row_idx+self.zeilen_counter, 2, QTableWidgetItem(instanz))  # Spalte 3 ist die Instanz
            for col_idx, item in enumerate(row, start=3):
                if isinstance(item, Decimal):
                    item = str(item)
                elif isinstance(item, int):
                    item = str(item)
                self.table_widget.setItem(row_idx+self.zeilen_counter, col_idx, QTableWidgetItem(item))

        #Zeilen-Counter auf aktuellen Wert setzen für die nächste Instanz
        self.zeilen_counter = self.zeilen_counter+row_idx+1

        conn.close()




class SqlWindow(QMainWindow):

    def __init__(self, main_instance):
        # Das Eingabewindow für das SQL-Komando wurde hier von Hand erstellt nicht über
        # den QT-Designer. Dies müsste man noch mal ändern
        super().__init__()
        self.setWindowTitle("SQL-Eingabe")
        self.main_instance = main_instance  # Speichern der Instanz der anderen Klasse

        # Haupt-Widget und Layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Erstelle ein QPlainTextEdit-Widget für da SQL-Statement
        self.plain_text_edit = QPlainTextEdit()
        layout.addWidget(self.plain_text_edit)

        # Erstelle ein QPlainTextEdit-Widget für die SQL-Beschreibung
        self.commend_text_edit = QPlainTextEdit()
        layout.addWidget(self.commend_text_edit)

        # Erstelle einen OK-Knopf
        self.ok_button = QPushButton("OK")
        layout.addWidget(self.ok_button)

        # Erstelle einen History lesen-Knopf
        self.history_button = QPushButton("History lesen")
        layout.addWidget(self.history_button)

        # Erstelle einen History schreiben-Knopf
        self.history_schreiben_button = QPushButton("History schreiben")
        layout.addWidget(self.history_schreiben_button)



        # Verbinde den OK-Knopf mit der Methode ok_button_clicked
        self.ok_button.clicked.connect(self.ok_button_clicked)

        # Verbinde den History-Knopf mit der Methode history_clicked
        self.history_button.clicked.connect(self.history_button_clicked)
        self.history_schreiben_button.clicked.connect(self.history_schreiben_button_clicked)

        # Letzte SQL Befehl und Beschreibung vorbelegen
        self.plain_text_edit.setPlainText(sql_input)
        self.commend_text_edit.setPlainText('')



    def ok_button_clicked(self):
        global sql_input

        sql_statement = self.plain_text_edit.toPlainText()


        sql_input = sql_statement
        print(f'OK button clicked. SQL-Statement: {sql_statement}')
        # Sie können hier weitere Aktionen hinzufügen, z.B. das Fenster schließen oder das SQL-Statement verarbeiten

        self.close()


        # Aufrufen der Methode aus der anderen Klasse
        self.main_instance.abfrage()

    def history_button_clicked(self):
        """Holt ein Eintrag aus der History-Tabelle"""

        global history_index

        print("History geklicked")

        print(f'Index {history_index}')
        history_json = read_json_history()

        # Neues Statement abrufen
        sql_statement = history_json[history_index]["sql"]
        print(f"Nächstes SQL-Statement: {sql_statement}")
        self.plain_text_edit.setPlainText(sql_statement)
        history_index += 1
        # Vermeiden von Indexfehlern (zirkuläre Navigation)
        if history_index >= len(history_json):
            history_index = 0  # Zurück zum Anfang

    def history_schreiben_button_clicked(self):
        """Ein SQL-Statement in die History schreiben"""

        print("History schreiben geklicked")
        # Wenn ein neues SQL-statement eingegeben wurde, dann in History ablegen.
        sql_statement = self.plain_text_edit.toPlainText()
        sql_commend = self.commend_text_edit.toPlainText()

        #wenn das sql_statement noch nicht in der History-Datei ist, dann das Statement mit
        #Kommentar abspeichern
        if is_sql_in_history(sql_statement):
            print(f'SQL-Befehl {sql_statement} schon in der History vorhanden')
        else:
            print(f'SQL-Befehl {sql_statement} mit Kommentar {sql_commend} in SQL-History ablegen')
            save_to_json_history(sql_statement, sql_commend)


class HistoryWindow(QMainWindow, Ui_Form):

    def __init__(self, main_instance):

        super().__init__()
        self.ui = Ui_Form()  # Instanz der UI-Klasse erstellen
        self.ui.setupUi(self)  # WICHTIG! Lädt die UI-Elemente
        self.setWindowTitle("SQL-History")
        self.main_instance = main_instance  # Speichern der Instanz der anderen Klasse
        self.setFixedSize(600, 500)  # Setzt eine feste Fenstergröße

        # Setze die Spaltenüberschriften
        self.ui.tableWidget_History.setHorizontalHeaderLabels(["Beschreibung", "SQL-Kommando"])

        # Button-Event verbinden
        self.ui.pushButton_ok.clicked.connect(self.use_selected_sql)

        # Event für Listenauswahl
        self.ui.tableWidget_History.itemClicked.connect(self.display_selected_sql)

        # Filter verbinden
        self.ui.lineEdit_filter.textChanged.connect(self.filter_table)

        # JSON-Datei laden
        self.load_sql_history()

    def filter_table(self):
        """Filtert die Commend-Spalte nach der im Filter eingegeben Text"""
        filter_text = self.ui.lineEdit_filter.text().lower()
        for row in range(self.ui.tableWidget_History.rowCount()):
            match = False
            for col in range(self.ui.tableWidget_History.columnCount()):
                item = self.ui.tableWidget_History.item(row, col)
                if item and filter_text in item.text().lower():
                    match = True
                    break
            self.ui.tableWidget_History.setRowHidden(row, not match)

    def load_sql_history(self):
        """Lädt die SQL-Historie aus der JSON-Datei und zeigt sie in der Liste an."""
        try:
            with open("sql_history.json", "r", encoding="utf-8") as file:
                history = json.load(file)


            self.ui.tableWidget_History.setRowCount(len(history))  # Anzahl der Zeilen setzen muss vor der for Schleife erfolgen
            for row, entry in enumerate(history):
                timestamp = entry["timestamp"]
                sql_query = entry["sql"]
                sql_commend = entry["commend"]

                # Setze die Werte in die Tabelle (Spalte 0 = Beschreibung, Spalte 1 = SQL)
                self.ui.tableWidget_History.setItem(row, 0, QTableWidgetItem(sql_commend))
                self.ui.tableWidget_History.setItem(row, 1, QTableWidgetItem(sql_query))



            # Optional: Spalten automatisch anpassen
            self.ui.tableWidget_History.resizeColumnsToContents()
            self.ui.tableWidget_History.resizeRowsToContents()

            # Speichern der Daten für Zugriff beim Klick
            self.history_data = history

        except Exception as e:
            QMessageBox.warning(self, "Fehler", "SQL-Historie konnte nicht geladen werden.")

    def display_selected_sql(self):
        """Zeigt das ausgewählte SQL-Statement im Textfeld an."""

        global sql_input, sql_commend

        selected_row = self.ui.tableWidget_History.currentRow()
        if selected_row >= 0:

            # Ausgewältes SQL-Statement in die globale Variable schreiben
            sql_input = self.history_data[selected_row]["sql"]
            sql_commend = self.history_data[selected_row]["commend"]

            self.ui.textEdit_SQL.setPlainText(self.history_data[selected_row]["sql"])

    def use_selected_sql(self):
        """Übernimmt das SQL-Statement für die Hauptanwendung."""
        global sql_input

        selected_sql = self.ui.textEdit_SQL.toPlainText()
        sql_input = selected_sql
        if selected_sql:
            self.close()
            self.main_instance.abfrage()




    def ok_button_clicked(self):
        global sql_input

        sql_statement = self.plain_text_edit.toPlainText()


        sql_input = sql_statement
        print(f'OK button clicked. SQL-Statement: {sql_statement}')
        # Sie können hier weitere Aktionen hinzufügen, z.B. das Fenster schließen oder das SQL-Statement verarbeiten

        self.close()


        # Aufrufen der Methode aus der anderen Klasse
        self.main_instance.abfrage()

    def history_button_clicked(self):
        # Ein SQL-Statement aus der History holen

        global history_index

        print("History geklicked")

        print(f'Index {history_index}')
        history_json = read_json_history()

        # Neues Statement abrufen
        sql_statement = history_json[history_index]["sql"]
        print(f"Nächstes SQL-Statement: {sql_statement}")
        self.plain_text_edit.setPlainText(sql_statement)
        history_index += 1
        # Vermeiden von Indexfehlern (zirkuläre Navigation)
        if history_index >= len(history_json):
            history_index = 0  # Zurück zum Anfang

    def history_schreiben_button_clicked(self):
        # Ein SQL-Statement aus in die History schreiben

        print("History schreiben geklicked")
        # Wenn ein neues SQL-statement eingegeben wurde, dann in History ablegen.
        sql_statement = self.plain_text_edit.toPlainText()
        if sql_input != sql_statement:
           print(f'statement {sql_statement} in SQL-History ablegen')
           save_to_json_history(sql_statement)



def update_sql_with_paths(sql, instanz):
    """Ergänzt Tabellen durch ihre Pfade Instanz(COMV7 oder DATV7)"""

    # SQL in Kleinbuchstaben umwandeln, um alle Varianten sicher zu erfassen
    #sql = sql.lower()

    # Wandelt in SQL-Stament nur from oder join in Kleinbuchstaben. Der
    # Rest bleibt so wie er erfasst wurde. Dies ist wichtig, für die
    # where Bedingungen. Hies ist Groß/Kleinschreibung wichtig
    pattern = r'\b(FROM|JOIN)\b'  # Nur 'FROM' und 'JOIN' als ganze Wörter finden
    normalized_sql = re.sub(pattern, lambda match: match.group(0).lower(), sql, flags=re.IGNORECASE)

    # Sucht alle Tabellennamen nach FROM oder JOIN
    matches = re.findall(r'\b(?:from|join)\s+(\w+)', normalized_sql)

    # Falls keine Tabellen gefunden wurden, gib das Original zurück
    if not matches:
        print(f'Keine Tabellen gefunden normalized_sql {normalized_sql}')
        return normalized_sql

    # Erstelle eine Ersetzungsliste für alle gefundenen Tabellen
    replacements = {}
    for table in matches:
        #path = ("/comv7" if table[-1] in ["2", "3"] else "datv7") + f".{table}"
        path = f'{instanz}comv7.{table}' if table[-1] in ["2", "3"] else f'{instanz}datv7.{table}'

        replacements[table] = path

    # **Tabellen direkt im SQL-String ersetzen**
    # for table, path in replacements.items():
    #     normalized_sql = normalized_sql.replace(f"from {table}", f"from {path}")
    #     normalized_sql = normalized_sql.replace(f"join {table}", f"join {path}")

    # mit re.sub alle tabellennamen, durch path mit tabellennamen ersetzen.
    # Hiermit werden auch bei mehrfach vorkommende Tabellen die Pfade mit reingenommen.
    for table, path in replacements.items():
        normalized_sql = re.sub(rf'\b{table}\b', path, normalized_sql, flags=re.IGNORECASE)

    #print(f'SQL-Statement mit den Pfaden {normalized_sql}')
    return normalized_sql



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.setWindowTitle(f'SQL-Master Version {version}')
    window.show()
    sys.exit(app.exec())



